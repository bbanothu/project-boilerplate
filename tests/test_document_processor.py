import pytest
import os
import tempfile
from pathlib import Path
from openpyxl import Workbook
from app.services.document_processor import process_documents, extract_text_from_xlsx, extract_text_from_pdf

@pytest.fixture
def temp_files(tmp_path):
    # Create temp PDF
    pdf_path = tmp_path / 'test.pdf'
    with open(pdf_path, 'w') as f:
        f.write('Sample PDF text')

    # Create temp XLSX
    xlsx_path = tmp_path / 'test.xlsx'
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'Sample'
    ws['B1'] = 'Data'
    wb.save(xlsx_path)

    return [str(pdf_path), str(xlsx_path)]

def test_process_documents(temp_files):
    result = process_documents(temp_files)
    assert 'document_0.pdf' in result
    assert 'document_1.xlsx' in result
    assert 'Sample PDF text' in result['document_0.pdf']
    assert 'Sample\tData' in result['document_1.xlsx']

def test_extract_text_from_xlsx():
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'Test'
        wb.save(tmp.name)
    text = extract_text_from_xlsx(tmp.name)
    assert 'Sheet1' in text
    assert 'Test' in text
    os.unlink(tmp.name)
