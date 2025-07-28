import os
from app.utils.pdf_utils import extract_text_from_pdf
from openpyxl import load_workbook


def extract_text_from_xlsx(file_path: str) -> str:
    """
    Extract text from an XLSX file.

    Args:
        file_path: Path to the XLSX file

    Returns:
        str: Extracted text from the XLSX file
    """
    text = ""
    wb = load_workbook(file_path, read_only=True, data_only=True)
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        text += f"Sheet: {sheet_name}\n"
        for row in sheet.iter_rows(values_only=True):
            row_text = "\t".join(str(cell) if cell is not None else "" for cell in row)
            if row_text.strip():
                text += row_text + "\n"
        text += "\n"
    return text


def process_documents(file_paths):
    """
    Process different types of documents and extract relevant information.

    Args:
        file_paths: List of paths to the documents

    Returns:
        dict: Extracted data from documents
    """
    extracted_data = {}

    for i, file_path in enumerate(file_paths):
        ext = os.path.splitext(file_path)[1].lower()
        key = f"document_{i}{ext}"
        if ext == ".pdf":
            extracted_data[key] = extract_text_from_pdf(file_path)
        elif ext == ".xlsx":
            extracted_data[key] = extract_text_from_xlsx(file_path)

    return extracted_data 